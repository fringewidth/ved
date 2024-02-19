import random
import numpy as np
import requests
import json
import os
import openpyxl
import re
import time

def askMistral(prompt):
    response = requests.post(
  url="https://openrouter.ai/api/v1/chat/completions",
  headers={
    "Authorization": f"Bearer {os.environ['OPEN_ROUNTER_KEY']}",
  },
  data=json.dumps({
    "model": "mistralai/mistral-7b-instruct:free", 
    "messages": [
      {"role": "user", "content": prompt}
    ]
  })
)
    try:
      return response.json()["choices"][0]["message"]["content"]
    except KeyError:
       return 0


def generate_paper(project_title, field):
  interval = 1
  while True:
    time.sleep(interval)
    ans = askMistral(f"Generate a single paper's title and abstract for a project entitled \"{project_title}\". Format it as Title: [title] Abstract: [Abstract]")
    if ans!=0:
       print('Error. Retrying...')
       interval *= 2
    else:
       break
  print(ans)
  title = re.findall(r'Title:(.*)\n?', ans)[0]
  abstract = re.findall(r'Abstract:\s+(.*)\n?', ans)[0]
  doi = generate_doi()
  url = generate_url(doi)
  citations = generate_citations(title)
  return [project_title, title, abstract, doi, url, citations]

def generate_doi():
  return f"10.1111/{''.join([random.choice('0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ') for i in range(11)])}"

def generate_url(doi):
    return f"https://doi.org/{doi}"

def generate_citations(title):
   return max(0, int(np.random.normal(loc=150, scale=100)) + 5*len(title) + np.random.randint(10, 50) * np.random.randint(-1,1))

srcfilePath = "D:\projects.xlsx"
src_ws = openpyxl.load_workbook(srcfilePath).active
destfilePath = "D:\papers.xlsx"
dest_wb = openpyxl.Workbook()
dest_ws = dest_wb.active

for i in range(26, src_ws.max_row+1):
   proj_title = src_ws.cell(row=i, column=1).value
   proj_field = src_ws.cell(row=i, column=3).value
   print(f"Generating for project {proj_title}...")
   for j in range(random.randint(1,3)):
      print(f"\tGenerating paper {j+1}...")
      ws_row = generate_paper(proj_title, proj_field)
      if not ws_row:
          continue
      print(f"\tGeneration Complete. Paper entitled {ws_row[1]} generated.")
      dest_ws.append(ws_row)
      dest_wb.save(destfilePath)